"""
Excel Work Order Generator
===========================
Generates a formatted Technical Work Order (Технічне Завдання) as an Excel file
using openpyxl.

Returns raw bytes that can be streamed as a file download from Streamlit.
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any

import openpyxl
from utils.logger import get_logger

logger = get_logger(__name__)
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


# Colour palette
_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_SUBHEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
_ALT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
_ACCENT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_TOTAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_SECTION_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_WHITE_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_BOLD_FONT = Font(name="Calibri", bold=True, size=11)
_NORMAL_FONT = Font(name="Calibri", size=10)
_SMALL_FONT = Font(name="Calibri", size=9)
_TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F4E79")
_ITALIC_FONT = Font(name="Calibri", italic=True, size=9, color="7F7F7F")


def _cell(ws, row: int, col: int, value: Any, font=None, fill=None,
          alignment=None, border=None) -> None:
    c = ws.cell(row=row, column=col, value=value)
    if font:
        c.font = font
    if fill:
        c.fill = fill
    if alignment:
        c.alignment = alignment
    if border:
        c.border = border


# ─────────────────────────────────────────────────────────────────────────────
# Sheet: Work Order per component
# ─────────────────────────────────────────────────────────────────────────────

_SECTION_MAP: dict[str, str] = {
    "prepress":             "Ручні роботи / Препрес",
    "roll_slitting":        "Порізка / Формування паперу",
    "sheet_format_cutting": "Порізка / Формування паперу",
    "card_cutting":         "Порізка / Формування паперу",
    "offset_printing":      "Друк",
    "digital_printing":     "Друк",
    "lamination":           "Ламінація",
    "uv_varnishing":        "Ламінація",
    "hot_foil_stamping":    "Ламінація",
    "chipboard_laminating": "Кашировка",
    "die_cutting":          "Висікання",
    "creasing":             "Рицовка",
    "quality_control":      "Ручні роботи / Пакування",
    "game_kit_assembly":    "Ручні роботи / Пакування",
    "shrink_wrapping":      "Ручні роботи / Пакування",
    "shipper_packing":      "Ручні роботи / Пакування",
    "palletizing":          "Ручні роботи / Пакування",
}
_PACKING_SECTION = "Ручні роботи / Пакування"


def _section_for_op(op_id: str) -> str:
    return _SECTION_MAP.get(op_id, _PACKING_SECTION)


def _narid_sheet_name(component_name: str) -> str:
    clean = component_name[:18].strip()
    return f"Наряд - {clean}"


def _add_narid_sheet(
    wb: openpyxl.Workbook,
    route: dict,
    work_order: dict,
    requirements: dict,
) -> None:
    comp_name = route.get("component_name", route.get("component_id", "Компонент"))
    sheet_name = _narid_sheet_name(comp_name)
    ws = wb.create_sheet(sheet_name)

    # Column widths
    for col, w in zip("ABCDEFGHIJ", [14, 10, 28, 14, 30, 22, 14, 10, 12, 12]):
        ws.column_dimensions[get_column_letter(ord(col) - ord("A") + 1)].width = w

    order_num  = work_order.get("order_number", "N/A")
    now_str    = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
    product    = work_order.get("product") or requirements.get("product_name", "—")
    client     = work_order.get("client") or requirements.get("client_name", "—")
    quantity   = requirements.get("quantity", "—")
    deadline   = requirements.get("deadline_days")
    mat        = route.get("material", {})
    mat_str    = "; ".join(f"{k}: {v}" for k, v in mat.items()) if mat else "—"

    r = 1
    # Header
    ws.merge_cells(f"A{r}:J{r}")
    _cell(ws, r, 1, f"Наряд № {order_num} від {now_str}",
          font=_BOLD_FONT, fill=_SUBHEADER_FILL,
          alignment=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[r].height = 20
    r += 1

    for label, val in [
        ("Назва", product),
        ("Контрагент", client),
        ("Тираж, од:", quantity),
    ]:
        ws.merge_cells(f"A{r}:C{r}")
        _cell(ws, r, 1, label, font=_BOLD_FONT, fill=_ALT_FILL, border=_THIN_BORDER)
        ws.merge_cells(f"D{r}:J{r}")
        _cell(ws, r, 4, val, font=_NORMAL_FONT, fill=_WHITE_FILL, border=_THIN_BORDER)
        r += 1

    if deadline:
        ws.merge_cells(f"A{r}:C{r}")
        _cell(ws, r, 1, "Планований термін:", font=_BOLD_FONT, fill=_ALT_FILL, border=_THIN_BORDER)
        ws.merge_cells(f"D{r}:J{r}")
        _cell(ws, r, 4, f"{deadline} днів", font=_NORMAL_FONT, fill=_WHITE_FILL, border=_THIN_BORDER)
        r += 1

    r += 1
    ws.merge_cells(f"A{r}:J{r}")
    _cell(ws, r, 1,
          "Нижче — зведення маршруту з системи Dyz-Art. "
          "Колонки «Формат мм», «Нето/брутто арк», «Штамп/код ШФ» тощо "
          "з ERP-наряду заповнюються вручну за потреби.",
          font=_ITALIC_FONT, alignment=Alignment(wrap_text=True))
    ws.row_dimensions[r].height = 28
    r += 1

    if mat:
        ws.merge_cells(f"A{r}:J{r}")
        _cell(ws, r, 1, f"Матеріали: {mat_str}",
              font=_SMALL_FONT, fill=_ACCENT_FILL,
              alignment=Alignment(horizontal="left"))
        r += 1

    r += 1

    # Group operations by section, preserving step order so the first
    # section reflects the earliest step (prepress), not whichever section
    # the LLM happened to emit first.
    ops = sorted(route.get("operations", []), key=lambda o: o.get("step", 0))
    sections: dict[str, list[dict]] = {}
    for op in ops:
        sec = _section_for_op(op.get("operation_id", ""))
        sections.setdefault(sec, []).append(op)

    op_counter = 1
    for section, section_ops in sections.items():
        # Section header
        ws.merge_cells(f"A{r}:J{r}")
        _cell(ws, r, 1, section,
              font=Font(name="Calibri", bold=True, size=10),
              fill=_SECTION_FILL,
              alignment=Alignment(horizontal="left", vertical="center"))
        ws.row_dimensions[r].height = 18
        r += 1

        # Column headers
        col_headers = ["Операція", "Елемент", "Виріб", "Тираж", "Назва операції",
                       "Обладнання", "Параметри", "Норма", "Примітки", "ТЗ"]
        for ci, h in enumerate(col_headers, start=1):
            _cell(ws, r, ci, h,
                  font=_WHITE_FONT, fill=_HEADER_FILL, border=_THIN_BORDER,
                  alignment=Alignment(horizontal="center", vertical="center"))
        ws.row_dimensions[r].height = 16
        r += 1

        for op in section_ops:
            params = op.get("parameters") or {}
            params_str = ", ".join(f"{k}: {v}" for k, v in params.items()) if params else "—"
            row_data = [
                f"Опер{op_counter}",
                None,
                comp_name,
                quantity,
                op.get("operation_name", op.get("operation_id", "")),
                op.get("machine") or "—",
                params_str,
                None,
                op.get("notes") or "—",
                "—",
            ]
            fill = _ALT_FILL if op_counter % 2 == 0 else _WHITE_FILL
            for ci, val in enumerate(row_data, start=1):
                _cell(ws, r, ci, val,
                      font=_NORMAL_FONT, fill=fill, border=_THIN_BORDER,
                      alignment=Alignment(vertical="center", wrap_text=True))
            ws.row_dimensions[r].height = 16
            r += 1
            op_counter += 1

        r += 1

    # Comments
    r += 1
    ws.merge_cells(f"A{r}:J{r}")
    _cell(ws, r, 1, "Коментар", font=_BOLD_FONT, fill=_ALT_FILL)
    r += 1
    notes = work_order.get("special_notes") or requirements.get("notes") or requirements.get("game_components_notes") or ""
    ws.merge_cells(f"A{r}:J{r}")
    _cell(ws, r, 1, notes or " ",
          font=_NORMAL_FONT, alignment=Alignment(wrap_text=True))
    ws.row_dimensions[r].height = 30


# ─────────────────────────────────────────────────────────────────────────────
# Sheet: Cost Calculation
# ─────────────────────────────────────────────────────────────────────────────

def _fill_cost_sheet(
    ws,
    cost_estimates: dict,
    work_order: dict,
    requirements: dict,
) -> None:
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14

    order_num = work_order.get("order_number", "N/A")
    date_str  = datetime.now().strftime("%d.%m.%Y")
    product   = work_order.get("product") or requirements.get("product_name", "—")
    client    = work_order.get("client") or requirements.get("client_name", "—")
    quantity  = cost_estimates.get("base_quantity", requirements.get("quantity", 0))
    margin    = cost_estimates.get("margin", 1.10)

    r = 1
    ws.merge_cells(f"A{r}:D{r}")
    _cell(ws, r, 1, f"КАЛЬКУЛЯЦІЯ  •  {order_num}  •  {date_str}",
          font=_TITLE_FONT, alignment=Alignment(horizontal="center"))
    ws.row_dimensions[r].height = 22
    r += 1

    for label, val in [("Замовник", client), ("Назва замовлення", product)]:
        _cell(ws, r, 1, label, font=_BOLD_FONT, fill=_ALT_FILL, border=_THIN_BORDER)
        ws.merge_cells(f"B{r}:D{r}")
        _cell(ws, r, 2, val, font=_NORMAL_FONT, fill=_WHITE_FILL, border=_THIN_BORDER)
        r += 1

    r += 1

    # Column headers
    col_headers = ["Стаття витрат", "Ставка / од.", f"Тираж {quantity} шт.", "Ціна за од."]
    for ci, h in enumerate(col_headers, start=1):
        _cell(ws, r, ci, h,
              font=_WHITE_FONT, fill=_HEADER_FILL, border=_THIN_BORDER,
              alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[r].height = 18
    r += 1

    breakdown = cost_estimates.get("breakdown", [])
    for comp in breakdown:
        comp_name = comp.get("component_name", "Компонент").upper()
        # Component section header
        ws.merge_cells(f"A{r}:D{r}")
        _cell(ws, r, 1, f"▶  {comp_name}",
              font=Font(name="Calibri", bold=True, size=11, color="FFFFFF"),
              fill=_SUBHEADER_FILL,
              alignment=Alignment(horizontal="left", vertical="center"))
        ws.row_dimensions[r].height = 20
        r += 1

        line_items = comp.get("line_items", [])
        for i, li in enumerate(line_items):
            total     = li.get("total", 0)
            per_unit  = round(total / quantity, 4) if quantity else 0
            fill      = _ALT_FILL if i % 2 == 0 else _WHITE_FILL
            _cell(ws, r, 1, li.get("item", ""),   font=_NORMAL_FONT, fill=fill, border=_THIN_BORDER)
            _cell(ws, r, 2, li.get("rate", ""),   font=_SMALL_FONT,  fill=fill, border=_THIN_BORDER)
            _cell(ws, r, 3, total,                font=_NORMAL_FONT, fill=fill, border=_THIN_BORDER,
                  alignment=Alignment(horizontal="right"))
            _cell(ws, r, 4, per_unit,             font=_NORMAL_FONT, fill=fill, border=_THIN_BORDER,
                  alignment=Alignment(horizontal="right"))
            r += 1

        subtotal    = comp.get("subtotal", 0)
        per_unit_st = round(subtotal / quantity, 4) if quantity else 0
        ws.merge_cells(f"A{r}:B{r}")
        _cell(ws, r, 1, f"Собівартість  «{comp.get('component_name', '')}»",
              font=_BOLD_FONT, fill=_TOTAL_FILL, border=_THIN_BORDER)
        _cell(ws, r, 3, subtotal,    font=_BOLD_FONT, fill=_TOTAL_FILL, border=_THIN_BORDER,
              alignment=Alignment(horizontal="right"))
        _cell(ws, r, 4, per_unit_st, font=_BOLD_FONT, fill=_TOTAL_FILL, border=_THIN_BORDER,
              alignment=Alignment(horizontal="right"))
        r += 1
        r += 1

    # Grand total
    total_cost    = cost_estimates.get("total_cost", 0)
    cost_per_unit = cost_estimates.get("cost_per_unit", 0)
    price_per_unit = cost_estimates.get("price_per_unit", 0)
    total_payment  = cost_estimates.get("total_payment", 0)
    margin_pct     = int(round((margin - 1) * 100))

    ws.merge_cells(f"A{r}:B{r}")
    _cell(ws, r, 1, "Собівартість всього",
          font=_BOLD_FONT, fill=_TOTAL_FILL, border=_THIN_BORDER)
    _cell(ws, r, 3, total_cost,    font=_BOLD_FONT, fill=_TOTAL_FILL, border=_THIN_BORDER,
          alignment=Alignment(horizontal="right"))
    _cell(ws, r, 4, cost_per_unit, font=_BOLD_FONT, fill=_TOTAL_FILL, border=_THIN_BORDER,
          alignment=Alignment(horizontal="right"))
    r += 1

    _cell(ws, r, 1, f"Рентабельність (+{margin_pct}%)", font=_NORMAL_FONT, border=_THIN_BORDER)
    _cell(ws, r, 2, f"×{margin}", font=_NORMAL_FONT, border=_THIN_BORDER)
    r += 1

    ws.merge_cells(f"A{r}:B{r}")
    _cell(ws, r, 1, "До оплати за одиницю", font=_BOLD_FONT, fill=_ACCENT_FILL, border=_THIN_BORDER)
    _cell(ws, r, 3, price_per_unit, font=_BOLD_FONT, fill=_ACCENT_FILL, border=_THIN_BORDER,
          alignment=Alignment(horizontal="right"))
    r += 1

    ws.merge_cells(f"A{r}:B{r}")
    _cell(ws, r, 1, f"До оплати (тираж {quantity} шт.)",
          font=_BOLD_FONT, fill=_ACCENT_FILL, border=_THIN_BORDER)
    _cell(ws, r, 3, total_payment, font=_BOLD_FONT, fill=_ACCENT_FILL, border=_THIN_BORDER,
          alignment=Alignment(horizontal="right"))
    r += 1

    tiers = cost_estimates.get("tiers", {})
    if tiers:
        r += 1
        ws.merge_cells(f"A{r}:D{r}")
        _cell(ws, r, 1, "Орієнтовна вартість для інших тиражів:",
              font=_BOLD_FONT, fill=_SECTION_FILL)
        r += 1
        for tier_label, tier_price in tiers.items():
            _cell(ws, r, 1, tier_label,  font=_NORMAL_FONT, border=_THIN_BORDER)
            _cell(ws, r, 3, tier_price,  font=_NORMAL_FONT, border=_THIN_BORDER,
                  alignment=Alignment(horizontal="right"))
            r += 1

    r += 1
    ws.merge_cells(f"A{r}:D{r}")
    _cell(ws, r, 1,
          "Орієнтовна вартість без ПДВ. Ставки наближені — "
          "для точного розрахунку уточнюйте у менеджера.",
          font=_ITALIC_FONT, alignment=Alignment(wrap_text=True))
    ws.row_dimensions[r].height = 24


# ─────────────────────────────────────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────────────────────────────────────

def generate_work_order_excel(
    work_order: dict,
    routes: list[dict],
    requirements: dict,
    cost_estimates: dict | None = None,
) -> bytes:
    """
    Generates a formatted Excel Work Order and returns it as bytes.

    Parameters
    ----------
    work_order     : structured work order dict from GenerationAgent
    routes         : production routes list from TechnologistAgent
    requirements   : client requirements dict
    cost_estimates : cost breakdown dict from calculate_costs (optional)

    Returns
    -------
    bytes — Excel file content
    """
    logger.info("Generating Excel work order")
    logger.debug(f"Work order keys: {list(work_order.keys())}")
    logger.debug(f"Routes count: {len(routes)}")

    wb = openpyxl.Workbook()

    # -----------------------------------------------------------------------
    # Sheet 1: Summary / Cover page
    # -----------------------------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Технічне Завдання"
    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 45

    row = 1
    ws_summary.merge_cells(f"A{row}:B{row}")
    _cell(ws_summary, row, 1, "ТЕХНІЧНЕ ЗАВДАННЯ НА ВИРОБНИЦТВО",
          font=_TITLE_FONT, alignment=Alignment(horizontal="center"))
    row += 1
    ws_summary.merge_cells(f"A{row}:B{row}")
    _cell(ws_summary, row, 1, "Dyz-Art | Виробничий відділ",
          font=Font(name="Calibri", italic=True, size=10, color="7F7F7F"),
          alignment=Alignment(horizontal="center"))
    row += 2

    meta = [
        ("Номер замовлення", work_order.get("order_number", "N/A")),
        ("Дата формування", datetime.now().strftime("%d.%m.%Y %H:%M")),
        ("Клієнт", work_order.get("client") or requirements.get("client_name", "—")),
        ("Назва продукту", work_order.get("product") or requirements.get("product_name", "—")),
        ("Тираж", f"{requirements.get('quantity', '—')} шт."),
        ("Дедлайн (днів)", f"{requirements.get('deadline_days', '—')} днів"),
        ("Мова продукту", requirements.get("language", "uk").upper()),
    ]

    for label, value in meta:
        _cell(ws_summary, row, 1, label, font=_BOLD_FONT, fill=_ALT_FILL, border=_THIN_BORDER,
              alignment=Alignment(vertical="center"))
        _cell(ws_summary, row, 2, str(value), font=_NORMAL_FONT, fill=_WHITE_FILL,
              border=_THIN_BORDER, alignment=Alignment(vertical="center"))
        row += 1

    row += 1
    notes = work_order.get("special_notes") or requirements.get("notes", "")
    if notes:
        ws_summary.merge_cells(f"A{row}:B{row}")
        _cell(ws_summary, row, 1, "⚠️ Примітки:", font=_BOLD_FONT)
        row += 1
        ws_summary.merge_cells(f"A{row}:B{row}")
        _cell(ws_summary, row, 1, notes, font=_NORMAL_FONT,
              alignment=Alignment(wrap_text=True))
        ws_summary.row_dimensions[row].height = 40

    # -----------------------------------------------------------------------
    # Sheet 2: Production Routes
    # -----------------------------------------------------------------------
    ws_routes = wb.create_sheet("Маршрути виробництва")
    ws_routes.column_dimensions["A"].width = 6
    ws_routes.column_dimensions["B"].width = 25
    ws_routes.column_dimensions["C"].width = 30
    ws_routes.column_dimensions["D"].width = 28
    ws_routes.column_dimensions["E"].width = 20
    ws_routes.column_dimensions["F"].width = 30

    r = 1
    for route in routes:
        comp_name = route.get("component_name", route.get("component_id", "Компонент"))
        duration = route.get("estimated_duration_hours", "—")

        ws_routes.merge_cells(f"A{r}:F{r}")
        _cell(ws_routes, r, 1, f"▶  {comp_name.upper()}  |  ~{duration} год",
              font=_WHITE_FONT, fill=_SUBHEADER_FILL,
              alignment=Alignment(horizontal="left", vertical="center"))
        ws_routes.row_dimensions[r].height = 22
        r += 1

        material = route.get("material", {})
        if material:
            ws_routes.merge_cells(f"A{r}:F{r}")
            mat_str = "  Матеріали: " + " | ".join(f"{k}: {v}" for k, v in material.items())
            _cell(ws_routes, r, 1, mat_str, font=Font(name="Calibri", italic=True, size=9),
                  fill=_ACCENT_FILL, alignment=Alignment(horizontal="left"))
            r += 1

        headers = ["№", "Операція", "Назва операції", "Обладнання", "Параметри", "Примітки"]
        for col_idx, h in enumerate(headers, start=1):
            _cell(ws_routes, r, col_idx, h,
                  font=_WHITE_FONT, fill=_HEADER_FILL, border=_THIN_BORDER,
                  alignment=Alignment(horizontal="center", vertical="center"))
        ws_routes.row_dimensions[r].height = 18
        r += 1

        sorted_ops = sorted(route.get("operations", []), key=lambda o: o.get("step", 0))
        for i, op in enumerate(sorted_ops):
            fill = _ALT_FILL if i % 2 == 0 else _WHITE_FILL
            params_str = ", ".join(
                f"{k}: {v}" for k, v in (op.get("parameters") or {}).items()
            )
            row_data = [
                op.get("step", i + 1),
                op.get("operation_id", ""),
                op.get("operation_name", ""),
                op.get("machine") or "—",
                params_str or "—",
                op.get("notes") or "—",
            ]
            for col_idx, val in enumerate(row_data, start=1):
                _cell(ws_routes, r, col_idx, val,
                      font=_NORMAL_FONT, fill=fill, border=_THIN_BORDER,
                      alignment=Alignment(vertical="center", wrap_text=True))
            ws_routes.row_dimensions[r].height = 16
            r += 1

        r += 1

    # -----------------------------------------------------------------------
    # Sheets 3..N: One narid per component
    # -----------------------------------------------------------------------
    for route in routes:
        _add_narid_sheet(wb, route, work_order, requirements)

    # -----------------------------------------------------------------------
    # Sheet last: Cost Calculation
    # -----------------------------------------------------------------------
    ws_cost = wb.create_sheet("Калькуляція")
    if cost_estimates:
        _fill_cost_sheet(ws_cost, cost_estimates, work_order, requirements)
    else:
        ws_cost.cell(row=1, column=1, value="Калькуляція буде доступна після розрахунку собівартості.")

    # -----------------------------------------------------------------------
    # Serialize to bytes
    # -----------------------------------------------------------------------
    logger.debug("Serializing workbook to bytes...")
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    excel_bytes = buffer.read()
    logger.info(f"Excel file generated: {len(excel_bytes)} bytes")
    return excel_bytes
